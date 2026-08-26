import hashlib
import hmac
import io
import logging
import os
import time
import traceback

import requests
from dotenv import load_dotenv
from fastapi import FastAPI, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pdfplumber import open as pdfplumber_open
from pandas import DataFrame, isna, notna, to_numeric

load_dotenv()

TURNSTILE_SITE_KEY = os.getenv("TURNSTILE_SITE_KEY")
TURNSTILE_SECRET_KEY = os.getenv("TURNSTILE_SECRET_KEY")
SESSION_SECRET = os.getenv("SESSION_SECRET") or os.urandom(32)
SESSION_MAX_AGE = 3600  # 1 jam

logging.basicConfig(
    filename="converter.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


def verify_turnstile(token):
    if not token:
        return False
    secret = TURNSTILE_SECRET_KEY
    if not secret:
        return False
    try:
        response = requests.post(
            "https://challenges.cloudflare.com/turnstile/v0/siteverify",
            data={
                "secret": secret,
                "response": token,
            },
        )
        result = response.json()
        return result.get("success", False)
    except Exception:
        return False


def _sign_session(data: str) -> str:
    key = SESSION_SECRET if isinstance(SESSION_SECRET, bytes) else SESSION_SECRET.encode()
    return hmac.new(key, data.encode(), hashlib.sha256).hexdigest()


def create_session() -> str:
    ts = str(int(time.time()))
    return f"{ts}.{_sign_session(ts)}"


def verify_session(token: str) -> bool:
    if not token or "." not in token:
        return False
    ts, sig = token.split(".", 1)
    if not hmac.compare_digest(_sign_session(ts), sig):
        return False
    try:
        return (int(time.time()) - int(ts)) < SESSION_MAX_AGE
    except ValueError:
        return False


def clean_table_data(table):
    if not table:
        return []
    cleaned = []
    for row in table:
        cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(cell for cell in cleaned_row):
            cleaned.append(cleaned_row)
    return cleaned


def clean_numeric_value(value):
    if isna(value) or str(value).strip() == "":
        return value
    s = str(value).strip()
    s = s.replace(".", "")
    return "".join(c for c in s if c.isdigit() or c == ",")


def clean_numeric_columns(df):
    if df.empty:
        return df
    target_cols = [9, 10, 11, 12, 19]
    for col_idx in target_cols:
        if col_idx < len(df.columns):
            col_name = df.columns[col_idx]
            df[col_name] = df[col_name].apply(clean_numeric_value)
            df[col_name] = df[col_name].str.replace(",", ".", regex=False)
            df[col_name] = to_numeric(df[col_name], errors="coerce")
    return df


def remove_footer_rows(df, remove_footer):
    if remove_footer:
        mask = ~df.apply(
            lambda row: any(
                "sipd-ri" in str(cell).lower() for cell in row if notna(cell)
            ),
            axis=1,
        )
        df = df[mask].reset_index(drop=True)
    return df


def format_excel(ws, auto_width):
    if auto_width:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def normalize_text(text):
    if text is None:
        return ""
    text = str(text).lower().strip()
    text = text.replace("\n", " ").replace("\r", " ")
    return " ".join(text.split())


def row_values_match_header(row, header_values):
    row_texts = [normalize_text(cell) for cell in row]
    header_texts = [normalize_text(cell) for cell in header_values]
    if len(row_texts) != len(header_texts):
        return False
    matches = sum(1 for a, b in zip(row_texts, header_texts) if a == b and a)
    return matches >= len(header_texts) * 0.7


def row_values_match_header_rows(row, header_rows):
    return any(row_values_match_header(row, hr) for hr in header_rows)


def is_numbering_row(row):
    cells = [str(cell).strip() for cell in row if cell is not None]
    if not cells:
        return False
    if not cells[0].isdigit():
        return False
    return all(c.replace(".", "").replace("-", "").isdigit() for c in cells)


def is_potential_parent(val):
    val = str(val).strip()
    if not val:
        return False
    if len(val) == 1 and val.isdigit():
        return True
    return "." in val


def get_parent_of(value, all_values):
    parts = value.split(".")
    for i in range(len(parts) - 1, 0, -1):
        candidate = ".".join(parts[:i])
        if candidate in all_values:
            return candidate
    return None


def find_direct_children(data_df):
    all_values = set()
    for idx in range(len(data_df)):
        if len(data_df.columns) > 1:
            val = str(data_df.iloc[idx, 1]).strip()
            if val:
                all_values.add(val)

    groups = []
    for idx in range(len(data_df)):
        if len(data_df.columns) > 1:
            val = str(data_df.iloc[idx, 1]).strip()
            if not val or not is_potential_parent(val):
                continue

            prefix = val + "."
            parent_segments = len(val.split("."))
            children = []

            for child_idx in range(idx + 1, len(data_df)):
                child_val = str(data_df.iloc[child_idx, 1]).strip()
                if not child_val or not is_potential_parent(child_val):
                    continue

                child_segments = len(child_val.split("."))
                if child_segments <= parent_segments:
                    break

                if child_val.startswith(prefix):
                    if get_parent_of(child_val, all_values) == val:
                        children.append(child_idx)

            if children:
                groups.append((idx, children))

    return groups


def apply_sum_formulas(ws, data_df, header_row_count):
    groups = find_direct_children(data_df)
    if not groups:
        return

    target_cols = [9, 10, 11, 12, 19]

    for parent_idx, children_indices in groups:
        parent_excel_row = parent_idx + header_row_count + 1
        child_rows = [child_idx + header_row_count + 1 for child_idx in children_indices]
        child_rows.sort()

        for col_idx in target_cols:
            if col_idx < len(data_df.columns):
                col_letter = get_column_letter(col_idx + 1)
                if len(child_rows) == 1:
                    formula = f"={col_letter}{child_rows[0]}"
                else:
                    child_refs = ",".join(f"{col_letter}{r}" for r in child_rows)
                    formula = f"=SUM({child_refs})"
                ws.cell(row=parent_excel_row, column=col_idx + 1, value=formula)


def convert_pdf_to_excel(pdf_file, remove_footer=True, auto_width=True):
    logger.info("Mulai konversi PDF")
    all_tables = []
    master_header_rows = None

    with pdfplumber_open(pdf_file) as pdf:
        total_pages = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            for table in tables:
                cleaned_table = clean_table_data(table)
                if not cleaned_table:
                    continue

                if master_header_rows is None:
                    master_header_rows = cleaned_table[:4]
                    all_tables.append(cleaned_table[4:])
                else:
                    first_rows = cleaned_table[:4]
                    is_repeated_header = any(
                        row_values_match_header(first_rows[j], master_header_rows[j])
                        for j in range(min(4, len(first_rows), len(master_header_rows)))
                    )
                    if is_repeated_header:
                        all_tables.append(cleaned_table[4:])
                    else:
                        all_tables.append(cleaned_table)

    if not all_tables:
        raise ValueError("Tidak ada tabel yang ditemukan di PDF ini.")

    flat_data = []
    for table in all_tables:
        flat_data.extend(table)

    if not flat_data and not master_header_rows:
        raise ValueError("Tidak ada data di PDF ini.")

    logger.info(f"PDF berhasil dibuka, total halaman: {total_pages}")

    header_df = DataFrame(master_header_rows)
    data_df = DataFrame(flat_data)

    logger.info(f"Data awal: {len(data_df)} baris")

    if not data_df.empty:
        data_df = data_df[
            ~data_df.apply(
                lambda row: row_values_match_header_rows(row.tolist(), master_header_rows),
                axis=1,
            )
        ].reset_index(drop=True)
        logger.info(f"Setelah filter header: {len(data_df)} baris")

    data_df = remove_footer_rows(data_df, remove_footer)
    data_df = data_df[
        ~data_df.apply(lambda row: is_numbering_row(row.tolist()), axis=1)
    ].reset_index(drop=True)
    data_df = data_df.dropna(how="all").reset_index(drop=True)
    data_df.columns = [str(c) for c in data_df.columns]
    data_df = data_df.loc[:, ~data_df.columns.str.contains("^Unnamed")]
    data_df = clean_numeric_columns(data_df)

    logger.info(f"Setelah filter footer/numbering/NaN: {len(data_df)} baris")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data PDF"

    for r_idx, row in enumerate(dataframe_to_rows(header_df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    start_row = len(master_header_rows) + 1
    for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=False), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    apply_sum_formulas(ws, data_df, len(master_header_rows))
    format_excel(ws, auto_width)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("Konversi selesai")
    return output, len(master_header_rows), len(data_df)


app = FastAPI(title="PDF ke Excel Converter - SIPD-RI")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/config")
def get_config():
    if not TURNSTILE_SITE_KEY or not TURNSTILE_SECRET_KEY:
        return JSONResponse(
            status_code=500,
            content={
                "error": "Konfigurasi Turnstile tidak ditemukan. Harap set TURNSTILE_SITE_KEY dan TURNSTILE_SECRET_KEY di file .env"
            },
        )
    return {"turnstile_site_key": TURNSTILE_SITE_KEY}


@app.post("/api/verify")
async def verify(token: str = Form(...)):
    if not TURNSTILE_SITE_KEY or not TURNSTILE_SECRET_KEY:
        raise HTTPException(
            status_code=500,
            detail="Konfigurasi Turnstile tidak ditemukan. Harap set TURNSTILE_SITE_KEY dan TURNSTILE_SECRET_KEY di file .env",
        )

    if not token or not verify_turnstile(token):
        raise HTTPException(
            status_code=403,
            detail="Verifikasi captcha gagal. Silakan coba lagi.",
        )

    return {"session": create_session()}


@app.post("/api/convert")
async def convert(
    session: str = Form(...),
    file: UploadFile = None,
    remove_footer: bool = Form(True),
    auto_width: bool = Form(True),
):
    if not TURNSTILE_SITE_KEY or not TURNSTILE_SECRET_KEY:
        raise HTTPException(
            status_code=500,
            detail="Konfigurasi Turnstile tidak ditemukan. Harap set TURNSTILE_SITE_KEY dan TURNSTILE_SECRET_KEY di file .env",
        )

    if not verify_session(session):
        raise HTTPException(
            status_code=403,
            detail="Sesi tidak valid. Silakan verifikasi captcha kembali.",
        )

    if file is None:
        raise HTTPException(status_code=400, detail="Silakan upload file PDF terlebih dahulu.")

    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File harus berformat PDF.")

    try:
        contents = await file.read()
        output, header_count, data_count = convert_pdf_to_excel(
            io.BytesIO(contents),
            remove_footer=remove_footer,
            auto_width=auto_width,
        )
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"Terjadi kesalahan: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_msg)

    excel_name = (file.filename or "converted").rsplit(".pdf", 1)[0] + ".xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{excel_name}"',
        "X-Header-Count": str(header_count),
        "X-Data-Count": str(data_count),
        "X-Filename": excel_name,
        "Access-Control-Expose-Headers": "X-Header-Count, X-Data-Count, X-Filename, Content-Disposition",
    }
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8501)
